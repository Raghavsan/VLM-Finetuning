"""
LLM-supervised invoice evaluation via OpenRouter.

Extends the heuristic evaluator by routing ambiguous comparisons (dates,
currencies, vendor/customer names) to an LLM that understands semantic
equivalence:
  - "26 February" == "26/02" == "2026-02-26"
  - "S$" == "SGD" == "Singapore Dollar"

Usage:
    export OPENROUTER_API_KEY="sk-or-..."
    python evaluate_llm_openrouter.py \
        --predictions data/company_dataset_copy/present_runs/predictions_base_updated_test \
        --gt          data/company_dataset_copy/test/GT \
        --out         data/eval/evaluation_report_llm.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import urllib.request
import urllib.error

try:
    from rapidfuzz import fuzz
except Exception:
    fuzz = None

# ── thresholds ────────────────────────────────────────────────────────────────
TEXT_CHAR_THRESHOLD = 0.90
TEXT_JACCARD_THRESHOLD = 0.75

# Fields the LLM will be asked about when heuristics disagree
LLM_CANDIDATE_FIELDS = {"invoice_date", "currency", "vendor_name", "customer_name"}

TARGET_FIELDS = {
    "invoice_number",
    "invoice_date",
    "vendor_name",
    "customer_name",
    "amount_before_vat",
    "vat_or_gst_amount",
    "amount_after_vat",
    "wht",
    "currency",
}

# ── OpenRouter config ─────────────────────────────────────────────────────────
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
DEFAULT_MODEL = "google/gemini-2.5-flash"
LLM_TIMEOUT = 30
LLM_MAX_RETRIES = 3
LLM_RETRY_DELAY = 2


# ── data structures ───────────────────────────────────────────────────────────
@dataclass(frozen=True)
class MatchedFiles:
    doc_id: str
    pred_path: Path
    gt_path: Path


@dataclass
class LLMStats:
    calls: int = 0
    hits: int = 0
    errors: int = 0
    cache_hits: int = 0
    total_tokens_in: int = 0
    total_tokens_out: int = 0
    _cache: dict = field(default_factory=dict, compare=False, repr=False)


_STATS = LLMStats()

CACHE_FILE = Path("data/eval/llm_cache.json")

def load_llm_cache() -> dict:
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"Warning: Failed to load LLM cache: {e}")
    return {}

def save_llm_cache(cache: dict) -> None:
    try:
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"Warning: Failed to save LLM cache: {e}")


def load_json(path: Path) -> Any:
    raw = path.read_text(encoding="utf-8-sig").strip()
    # Try parsing after stripping markdown fences
    try:
        cleaned = re.sub(r"^```(?:json)?\s*", "", raw)
        cleaned = re.sub(r"\s*```$", "", cleaned).strip()
        return json.loads(cleaned)
    except Exception:
        pass

    # Find the outermost curly braces if simple parsing fails
    match = re.search(r"(\{.*\})", raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1).strip())
        except Exception:
            pass

    # Fallback to raising standard exception
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw)
    cleaned = re.sub(r"\s*```$", "", cleaned).strip()
    return json.loads(cleaned)


def rows_from_json(data: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(data, dict):
        for key in ("metadata", "predictions", "fields", "results"):
            if isinstance(data.get(key), list):
                rows = [r for r in data[key] if isinstance(r, dict)]
                break
        else:
            if "field" in data and "value" in data:
                rows = [data]
            else:
                for k, v in data.items():
                    rows.append({"field": k, "value": v})
    elif isinstance(data, list):
        rows = [r for r in data if isinstance(r, dict)]

    return [r for r in rows if str(r.get("field", "")).strip() in TARGET_FIELDS]


# ── file matching ─────────────────────────────────────────────────────────────
def strip_upload_suffix(stem: str) -> str:
    stem = re.sub(r"_\d{10,}$", "", stem)
    stem = re.sub(r"_gt$", "", stem, flags=re.IGNORECASE)
    return stem.strip()


def normalize_doc_key(stem: str) -> str:
    return re.sub(r"\s+", " ", strip_upload_suffix(stem)).casefold().strip()


def build_key_index(paths: list[Path]) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = defaultdict(list)
    for path in paths:
        index[normalize_doc_key(path.stem)].append(path)
    return index


def select_unique(key: str, index: dict[str, list[Path]]) -> Path | None:
    matches = index.get(key, [])
    return matches[0] if len(matches) == 1 else None


def find_matches(
    pred_dir: Path, gt_dir: Path
) -> tuple[list[MatchedFiles], list[dict[str, Any]]]:
    gt_index = build_key_index(sorted(gt_dir.glob("*.json")))
    matches: list[MatchedFiles] = []
    unmatched: list[dict[str, Any]] = []
    for pred_path in sorted(pred_dir.glob("*.json")):
        key = normalize_doc_key(pred_path.stem)
        gt_path = select_unique(key, gt_index)
        if gt_path is None:
            unmatched.append(
                {
                    "prediction_file": str(pred_path),
                    "reason": "no unique GT match",
                    "normalized_key": key,
                    "candidate_count": len(gt_index.get(key, [])),
                }
            )
            continue
        matches.append(MatchedFiles(strip_upload_suffix(gt_path.stem), pred_path, gt_path))
    print(f"Matched: {len(matches)}  |  Unmatched: {len(unmatched)}")
    for u in unmatched:
        print(" ", u)
    return matches, unmatched


# ── numeric helpers ───────────────────────────────────────────────────────────
def normalize_number(value: str) -> str | None:
    s = re.sub(r"[^\d,.\-]", "", value)
    if not re.search(r"\d", s):
        return None
    if "," in s and "." in s:
        dec = "," if s.rfind(",") > s.rfind(".") else "."
        thou = "." if dec == "," else ","
        s = s.replace(thou, "").replace(dec, ".")
    elif "," in s:
        s = (
            s.replace(".", "").replace(",", ".")
            if re.search(r",\d{1,2}$", s)
            else s.replace(",", "")
        )
    elif s.count(".") > 1:
        parts = s.split(".")
        s = "".join(parts[:-1]) + "." + parts[-1]
    try:
        number = float(s)
    except ValueError:
        return None
    return f"{number:.2f}".rstrip("0").rstrip(".")


# ── date normalization (extended) ──────────────────────────────────────────────
_DATE_FORMATS = (
    "%d.%m.%Y", "%d.%m.%y",
    "%d/%m/%Y", "%d/%m/%y",
    "%m/%d/%Y", "%m/%d/%y",
    "%Y-%m-%d",
    "%d-%m-%Y", "%d-%m-%y",
    "%B %d, %Y", "%b %d, %Y",
    "%d %B %Y", "%d %b %Y",
    "%d %B, %Y", "%d %b, %Y",
    "%B %d %Y", "%b %d %Y",
    "%Y%m%d",
)


def normalize_date(value: str) -> str | None:
    s = value.strip()
    s_upper = s.upper()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
        try:
            return datetime.strptime(s_upper, fmt.upper()).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


# ── currency normalization ────────────────────────────────────────────────────
_CURRENCY_MAP: dict[str, str] = {
    # Singapore
    "sgd": "SGD", "s$": "SGD", "sg$": "SGD", "singapore dollar": "SGD",
    "singapore dollars": "SGD",
    # Malaysian
    "myr": "MYR", "rm": "MYR", "ringgit": "MYR", "malaysian ringgit": "MYR",
    # US
    "usd": "USD", "us$": "USD", "u.s.$": "USD", "dollar": "USD", "dollars": "USD",
    "$": "USD",
    # Euro
    "eur": "EUR", "\u20ac": "EUR", "euro": "EUR", "euros": "EUR",
    # British
    "gbp": "GBP", "\u00a3": "GBP", "pound": "GBP", "pounds": "GBP",
    # Australian
    "aud": "AUD", "a$": "AUD", "au$": "AUD", "australian dollar": "AUD",
    # Canadian
    "cad": "CAD", "c$": "CAD", "ca$": "CAD", "canadian dollar": "CAD",
    # Thai
    "thb": "THB", "\u0e3f": "THB", "baht": "THB",
    # Indonesian
    "idr": "IDR", "rp": "IDR", "rupiah": "IDR",
    # Japanese
    "jpy": "JPY", "\u00a5": "JPY", "yen": "JPY",
    # Chinese
    "cny": "CNY", "rmb": "CNY", "yuan": "CNY",
    # Indian
    "inr": "INR", "\u20b9": "INR", "rupee": "INR", "rupees": "INR",
    # Hong Kong
    "hkd": "HKD", "hk$": "HKD",
    # Swiss
    "chf": "CHF", "franc": "CHF", "francs": "CHF",
    # Korean
    "krw": "KRW", "\u20a9": "KRW", "won": "KRW",
    # Vietnamese
    "vnd": "VND", "\u20ab": "VND", "dong": "VND",
    # Philippine
    "php": "PHP", "\u20b1": "PHP", "peso": "PHP",
    # New Zealand
    "nzd": "NZD", "nz$": "NZD",
    # South African
    "zar": "ZAR", "rand": "ZAR",
    # Brazilian
    "brl": "BRL", "r$": "BRL", "real": "BRL",
    # Mexican
    "mxn": "MXN", "mex$": "MXN",
    # Turkish
    "try": "TRY", "\u20ba": "TRY", "lira": "TRY",
    # Swedish
    "sek": "SEK",
    # Norwegian
    "nok": "NOK",
    # Danish
    "dkk": "DKK",
    # Czech
    "czk": "CZK",
    # Polish
    "pln": "PLN", "z\u0142": "PLN", "zloty": "PLN",
    # Russian
    "rub": "RUB", "\u20bd": "RUB", "ruble": "RUB",
    # UAE
    "aed": "AED", "dirham": "AED",
    # Saudi
    "sar": "SAR", "riyal": "SAR",
}


def normalize_currency(value: Any) -> str:
    if value is None:
        return ""
    key = str(value).strip().casefold()
    return _CURRENCY_MAP.get(key, str(value).strip().upper())


# ── general normalization ─────────────────────────────────────────────────────
def normalize_value(value: Any, field: str = "") -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    field_l = field.casefold()

    if any(t in field_l for t in ("number", "invoice_no", "invoice_id", "id")):
        raw = re.sub(r"\.0$", "", raw)

    if "date" in field_l:
        nd = normalize_date(raw)
        if nd:
            return nd

    if "currency" in field_l:
        return normalize_currency(raw)

    if any(t in field_l for t in ("amount", "vat", "total", "tax", "price", "subtotal", "balance", "wht")):
        nn = normalize_number(raw)
        if nn is not None:
            return nn

    normalized = raw.casefold()
    normalized = re.sub(r"\b(inv(?:oice)?|rechnung|no|nr|number|date)\b[:#.]?", "", normalized)
    return re.sub(r"[\s\-_/.,:;]+", "", normalized)


# ── text similarity ───────────────────────────────────────────────────────────
def text_for_similarity(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold())


def char_similarity(a: Any, b: Any) -> float:
    aa = text_for_similarity(a)
    bb = text_for_similarity(b)
    if not aa and not bb:
        return 1.0
    if fuzz:
        return fuzz.ratio(aa, bb) / 100.0
    return SequenceMatcher(a=aa, b=bb).ratio()


def words(value: Any) -> list[str]:
    return re.findall(r"\w+", text_for_similarity(value))


def word_level_metrics(gt_text: Any, pred_text: Any) -> dict[str, float]:
    gt_words = words(gt_text)
    pred_words = words(pred_text)
    if not gt_words and not pred_words:
        return {"word_precision": 1.0, "word_recall": 1.0, "word_f1": 1.0,
                "word_match_ratio": 1.0, "word_jaccard": 1.0}
    if not gt_words:
        return {"word_precision": 0.0, "word_recall": 1.0, "word_f1": 0.0,
                "word_match_ratio": 0.0, "word_jaccard": 0.0}
    if not pred_words:
        return {"word_precision": 0.0, "word_recall": 0.0, "word_f1": 0.0,
                "word_match_ratio": 0.0, "word_jaccard": 0.0}

    gt_counts = Counter(gt_words)
    pred_counts = Counter(pred_words)
    overlap = sum(min(gt_counts[w], pred_counts[w]) for w in gt_counts)
    precision = overlap / len(pred_words)
    recall = overlap / len(gt_words)
    f1 = (2 * precision * recall / (precision + recall)) if precision + recall else 0.0
    sm = SequenceMatcher(a=gt_words, b=pred_words)
    matches_count = sum(block.size for block in sm.get_matching_blocks())
    wer = (max(len(gt_words), len(pred_words)) - matches_count) / len(gt_words)
    gt_set = set(gt_words)
    pred_set = set(pred_words)
    jaccard = len(gt_set & pred_set) / len(gt_set | pred_set) if gt_set or pred_set else 1.0
    return {
        "word_precision": round(precision, 4),
        "word_recall": round(recall, 4),
        "word_f1": round(f1, 4),
        "word_match_ratio": round(max(0.0, 1 - wer), 4),
        "word_jaccard": round(jaccard, 4),
    }


def field_kind(field: str) -> str:
    fl = field.casefold()
    if "date" in fl:
        return "date"
    if any(t in fl for t in ("amount", "vat", "total", "tax", "price", "subtotal", "balance", "wht")):
        return "numeric"
    if any(t in fl for t in ("number", "invoice_no", "invoice_id", " id", "_id", "code", "iban", "swift", "currency")):
        return "identifier"
    return "text"


# ── heuristic value status ────────────────────────────────────────────────────
def value_status_heuristic(field: str, gt_value: Any, pred_value: Any) -> tuple[str, str]:
    kind = field_kind(field)
    if normalize_value(gt_value, field) == normalize_value(pred_value, field):
        return "correct", kind
    if kind in {"numeric", "date", "identifier"}:
        return "incorrect", kind
    c_sim = char_similarity(gt_value, pred_value)
    jaccard = word_level_metrics(gt_value, pred_value)["word_jaccard"]
    if c_sim >= TEXT_CHAR_THRESHOLD or jaccard >= TEXT_JACCARD_THRESHOLD:
        return "correct_fuzzy", kind
    return "incorrect", kind


# ── LLM judge ─────────────────────────────────────────────────────────────────
_LLM_SYSTEM_PROMPT = """\
You are a precise invoice data evaluator. You will be given a field name, a ground-truth value, and a predicted value from an invoice extraction model.

Your task: decide whether the prediction is SEMANTICALLY EQUIVALENT to the ground truth, even if the textual form differs.

Rules:
- Dates: "26 February 2026", "26/02/2026", "2/26/2026", "Feb 26, 2026", "2026-02-26" are all equivalent.
- Currencies: "S$", "SGD", "Singapore Dollar(s)" are equivalent. "MYR", "RM", "Ringgit" are equivalent. Apply similar logic for all currencies.
- Company/vendor names: minor abbreviations, punctuation differences, or commonly dropped words (Pte, Ltd, Inc, Co) should be treated as equivalent. Spelling mistakes that clearly refer to the same entity are OK.
- Numbers: "1,500.00", "1500", "1 500" are equivalent.
- Do NOT be lenient about fundamentally different values (different date, different amount, completely different company name, different currency).

Respond with ONLY a JSON object on one line:
{"equivalent": true}  or  {"equivalent": false, "reason": "short reason"}
"""


def _llm_cache_key(field: str, gt: Any, pred: Any) -> str:
    return json.dumps([field, str(gt), str(pred)], ensure_ascii=False)


def ask_llm(
    field: str,
    gt_value: Any,
    pred_value: Any,
    api_key: str,
    model: str = DEFAULT_MODEL,
) -> bool | None:
    cache_key = _llm_cache_key(field, gt_value, pred_value)
    if cache_key in _STATS._cache:
        _STATS.cache_hits += 1
        return _STATS._cache[cache_key]

    user_msg = (
        f'Field: "{field}"\n'
        f'Ground truth: "{gt_value}"\n'
        f'Prediction:   "{pred_value}"\n'
        "Are these semantically equivalent for this invoice field?"
    )
    payload = json.dumps(
        {
            "model": model,
            "messages": [
                {"role": "system", "content": _LLM_SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
            ],
            "temperature": 0.0,
            "max_tokens": 80,
        }
    ).encode("utf-8")

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://github.com/alltius/invoice-eval",
        "X-Title": "Invoice Evaluation",
    }

    for attempt in range(LLM_MAX_RETRIES):
        try:
            req = urllib.request.Request(
                OPENROUTER_API_URL,
                data=payload,
                headers=headers,
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=LLM_TIMEOUT) as resp:
                body = json.loads(resp.read().decode("utf-8"))

            content = body["choices"][0]["message"]["content"].strip()
            content = re.sub(r"^```(?:json)?\s*", "", content)
            content = re.sub(r"\s*```$", "", content).strip()
            result_obj = json.loads(content)
            equivalent = bool(result_obj.get("equivalent", False))

            usage = body.get("usage", {})
            _STATS.total_tokens_in += usage.get("prompt_tokens", 0)
            _STATS.total_tokens_out += usage.get("completion_tokens", 0)
            _STATS.calls += 1
            _STATS._cache[cache_key] = equivalent
            return equivalent

        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            print(f"  [LLM] HTTP {e.code} on attempt {attempt+1}: {err_body[:200]}")
            if e.code in (429, 503):
                time.sleep(LLM_RETRY_DELAY * (attempt + 1))
            else:
                _STATS.errors += 1
                return None

        except Exception as exc:
            print(f"  [LLM] Error on attempt {attempt+1}: {exc}")
            time.sleep(LLM_RETRY_DELAY)

    _STATS.errors += 1
    return None


# ── combined value status (heuristic + LLM) ───────────────────────────────────
def value_status(
    field: str,
    gt_value: Any,
    pred_value: Any,
    api_key: str | None,
    model: str,
) -> tuple[str, str]:
    heuristic_status, kind = value_status_heuristic(field, gt_value, pred_value)

    if heuristic_status != "incorrect":
        return heuristic_status, kind

    if api_key and field in LLM_CANDIDATE_FIELDS:
        if (gt_value is not None and str(gt_value).strip()
                and pred_value is not None and str(pred_value).strip()):
            equivalent = ask_llm(field, gt_value, pred_value, api_key, model)
            if equivalent is True:
                _STATS.hits += 1
                return "correct_llm", kind

    return "incorrect", kind


# ── amount pairing ────────────────────────────────────────────────────────────
def comparable_number(value: Any) -> float | None:
    normalized = normalize_number(str(value))
    if normalized is None:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def amount_pair_score(gt: dict[str, Any], pred: dict[str, Any], field: str) -> tuple[float, float, float]:
    gt_num = comparable_number(gt.get("value"))
    pred_num = comparable_number(pred.get("value"))
    if gt_num is not None and pred_num is not None:
        diff = abs(gt_num - pred_num)
        scale = max(abs(gt_num), abs(pred_num), 1.0)
        numeric_score = max(0.0, 1.0 - diff / scale)
    else:
        numeric_score = 0.0
    exact = 1.0 if normalize_value(gt.get("value"), field) == normalize_value(pred.get("value"), field) else 0.0
    return (exact * 3.0 + numeric_score * 2.0, numeric_score, exact)


def best_one_to_one_pairs(
    gt_rows: list[dict[str, Any]],
    pred_rows: list[dict[str, Any]],
    field: str,
) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    if not gt_rows or not pred_rows:
        return []
    if len(gt_rows) == 1:
        return [(gt_rows[0], max(pred_rows, key=lambda p: amount_pair_score(gt_rows[0], p, field)))]

    pair_count = min(len(gt_rows), len(pred_rows))
    scores = {
        (gi, pi): amount_pair_score(gt, pred, field)[0]
        for gi, gt in enumerate(gt_rows)
        for pi, pred in enumerate(pred_rows)
    }
    best_score = -1.0
    best_pairs: list[tuple[int, int]] = []

    def search(gi: int, used: set, pairs: list, total: float) -> None:
        nonlocal best_score, best_pairs
        if len(pairs) == pair_count or gi == len(gt_rows):
            if total > best_score:
                best_score = total
                best_pairs = list(pairs)
            return
        if len(gt_rows) - gi > pair_count - len(pairs):
            search(gi + 1, used, pairs, total)
        for pi in range(len(pred_rows)):
            if pi in used:
                continue
            pairs.append((gi, pi))
            used.add(pi)
            search(gi + 1, used, pairs, total + scores[(gi, pi)])
            used.discard(pi)
            pairs.pop()

    search(0, set(), [], 0.0)
    return [(gt_rows[gi], pred_rows[pi]) for gi, pi in best_pairs]


def row_score(gt: dict[str, Any], pred: dict[str, Any], field: str) -> tuple[float, float]:
    exact = 1.0 if normalize_value(gt.get("value"), field) == normalize_value(pred.get("value"), field) else 0.0
    sim = char_similarity(gt.get("value"), pred.get("value"))
    return (exact * 2.0 + sim, sim)


# ── per-document evaluation ───────────────────────────────────────────────────
def evaluate_document(
    doc_id: str,
    gt_rows: list[dict[str, Any]],
    pred_rows: list[dict[str, Any]],
    api_key: str | None,
    model: str,
) -> list[dict[str, Any]]:
    gt_by_field: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in gt_rows:
        f = str(row.get("field", "")).strip()
        if f:
            gt_by_field[f].append(row)

    preds_by_field: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for pred in pred_rows:
        f = str(pred.get("field", "")).strip()
        if f:
            preds_by_field[f].append(pred)

    records: list[dict[str, Any]] = []

    def is_non_empty(val: Any) -> bool:
        return val is not None and str(val).strip() != ""

    def make_record(
        field: str,
        pred: dict[str, Any] | None,
        gt: dict[str, Any] | None,
        matching_note: str = "",
    ) -> dict[str, Any]:
        gt_value = gt.get("value") if gt else None
        pred_value = pred.get("value") if pred else None
        
        gt_exists = is_non_empty(gt_value)
        pred_exists = is_non_empty(pred_value)
        
        kind = field_kind(field)
        
        if gt_exists and pred_exists:
            status, _ = value_status(field, gt_value, pred_value, api_key, model)
        elif gt_exists and not pred_exists:
            status = "missing"
        elif not gt_exists and pred_exists:
            status = "hallucinated"
        else:
            status = "correct_empty"

        w_metrics = word_level_metrics(gt_value, pred_value)
        return {
            "doc_id": doc_id,
            "field_group": "sparse_predictions",
            "field": field,
            "field_kind": kind,
            "gt_value": gt_value,
            "pred_value": pred_value,
            "status": status,
            "counts_as_tp": gt_exists and pred_exists and (status in {"correct", "correct_fuzzy", "correct_llm"}),
            "matching_note": matching_note,
            "char_similarity": round(char_similarity(gt_value, pred_value), 4) if (gt_exists and pred_exists) else 0.0,
            "matched_gt_row_id": gt.get("row_id") if gt else None,
            **w_metrics,
        }

    for field in TARGET_FIELDS:
        candidates = gt_by_field.get(field, [])
        field_preds = preds_by_field.get(field, [])
        
        if not candidates and not field_preds:
            records.append(make_record(field, None, None))
            continue
            
        matched_gt_ids = set()
        
        if field_preds:
            if field == "amount_after_vat" and candidates:
                pairs = best_one_to_one_pairs(candidates, field_preds, field)
                note = (
                    "amount_one_gt_closest_pred"
                    if len(candidates) == 1 and len(field_preds) > 1
                    else "amount_one_to_one"
                )
                for gt_row, pred_row in pairs:
                    records.append(make_record(field, pred_row, gt_row, note))
                    if gt_row:
                        matched_gt_ids.add(id(gt_row))
            else:
                for pred in field_preds:
                    if candidates:
                        gt = max(candidates, key=lambda row: row_score(row, pred, field))
                        records.append(make_record(field, pred, gt))
                        if gt:
                            matched_gt_ids.add(id(gt))
                    else:
                        records.append(make_record(field, pred, None))
                        
        for gt in candidates:
            if id(gt) not in matched_gt_ids:
                if is_non_empty(gt.get("value")):
                    records.append(make_record(field, None, gt))

    return records


# ── aggregate metrics ─────────────────────────────────────────────────────────
def compute_prf1(records: list[dict[str, Any]]) -> dict[str, Any]:
    correct = 0
    incorrect = 0
    missing = 0
    hallucinated = 0
    correct_empty = 0

    for r in records:
        status = r["status"]
        if status in {"correct", "correct_fuzzy", "correct_llm"}:
            correct += 1
        elif status == "incorrect":
            incorrect += 1
        elif status == "missing":
            missing += 1
        elif status == "hallucinated":
            hallucinated += 1
        elif status == "correct_empty":
            correct_empty += 1

    total_gt_exists = correct + incorrect + missing
    total_pred_exists = correct + incorrect + hallucinated

    precision = correct / (correct + incorrect) if (correct + incorrect) > 0 else 0.0
    recall = (correct + incorrect) / total_gt_exists if total_gt_exists > 0 else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0.0
    hallucination_rate = hallucinated / total_pred_exists if total_pred_exists > 0 else 0.0

    n = len(records)
    field_accuracy = (correct + correct_empty) / n if n > 0 else 0.0
    
    correct_exact = sum(1 for r in records if r["status"] == "correct")
    exact_accuracy = (correct_exact + correct_empty) / n if n > 0 else 0.0

    return {
        "n_fields": n,
        "true_positive": correct,
        "incorrect": incorrect,
        "missing": missing,
        "hallucinated": hallucinated,
        "correct_empty": correct_empty,
        "total_gt_exists": total_gt_exists,
        "total_pred_exists": total_pred_exists,
        "correct_exact": correct_exact,
        "correct_fuzzy": sum(1 for r in records if r["status"] == "correct_fuzzy"),
        "correct_llm": sum(1 for r in records if r["status"] == "correct_llm"),
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1_score": round(f1, 4),
        "hallucination_rate": round(hallucination_rate, 4),
        "field_accuracy": round(field_accuracy, 4),
        "exact_accuracy": round(exact_accuracy, 4),
        "fuzzy_review_rate": round(sum(1 for r in records if r["status"] == "correct_fuzzy") / n, 4) if n > 0 else 0.0,
        "llm_rescue_rate": round(sum(1 for r in records if r["status"] == "correct_llm") / n, 4) if n > 0 else 0.0,
        "avg_word_precision": round(sum(r.get("word_precision", 0) for r in records) / n, 4) if n > 0 else 0.0,
        "avg_word_recall": round(sum(r.get("word_recall", 0) for r in records) / n, 4) if n > 0 else 0.0,
        "avg_word_f1": round(sum(r.get("word_f1", 0) for r in records) / n, 4) if n > 0 else 0.0,
        "avg_word_jaccard": round(sum(r.get("word_jaccard", 0) for r in records) / n, 4) if n > 0 else 0.0,
        "avg_char_similarity": round(sum(r.get("char_similarity", 0) for r in records) / n, 4) if n > 0 else 0.0,
    }


def breakdown(records: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in records:
        grouped[str(row.get(key, ""))].append(row)
    rows = []
    for name, recs in sorted(grouped.items()):
        item = compute_prf1(recs)
        item[key] = name
        rows.append(item)
    return rows


def error_analysis(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    errors = [r for r in records if r["status"] not in {"correct", "correct_fuzzy", "correct_llm", "correct_empty"}]
    near_misses = [r for r in errors if r["status"] == "incorrect" and r["char_similarity"] >= 0.85]
    errors_sorted = sorted(errors, key=lambda x: (x["status"] != "incorrect", x["char_similarity"]))
    return errors_sorted, near_misses


# ── Excel report ──────────────────────────────────────────────────────────────
def write_excel_report(
    out_path: Path,
    summary: dict[str, Any],
    group_rows: list[dict[str, Any]],
    field_rows: list[dict[str, Any]],
    doc_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    errors: list[dict[str, Any]],
    near_misses: list[dict[str, Any]],
    load_errors: list[dict[str, Any]],
    unmatched_predictions: list[dict[str, Any]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2F5496")
    fills = {
        "correct": PatternFill("solid", start_color="C6EFCE"),
        "correct_fuzzy": PatternFill("solid", start_color="D9EAF7"),
        "correct_llm": PatternFill("solid", start_color="B7E1CD"),
        "correct_empty": PatternFill("solid", start_color="E2EFDA"),
        "incorrect": PatternFill("solid", start_color="FFC7CE"),
        "missing": PatternFill("solid", start_color="FFEB9C"),
        "hallucinated": PatternFill("solid", start_color="D9D2E9"),
    }

    def write_table(ws, headers, rows, status_col=None):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
            if status_col and status_col in headers:
                fill = fills.get(str(row.get(status_col)))
                if fill:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=ws.max_row, column=col).fill = fill
        for col in range(1, len(headers) + 1):
            vals = [str(headers[col - 1])] + [str(row.get(headers[col - 1], "")) for row in rows]
            ws.column_dimensions[get_column_letter(col)].width = min(max(len(v) for v in vals) + 2, 60)
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Summary"
    write_table(ws, ["Metric", "Value"], [{"Metric": k, "Value": v} for k, v in summary.items()])

    metric_headers = [
        "n_fields", "true_positive", "incorrect", "missing", "hallucinated", "correct_empty",
        "total_gt_exists", "total_pred_exists",
        "precision", "recall", "f1_score", "hallucination_rate", "field_accuracy",
        "exact_accuracy", "fuzzy_review_rate", "llm_rescue_rate",
        "avg_word_precision", "avg_word_recall", "avg_word_f1", "avg_word_jaccard", "avg_char_similarity",
    ]
    ws2 = wb.create_sheet("By Field Group")
    write_table(ws2, ["field_group", *metric_headers], group_rows)
    ws3 = wb.create_sheet("By Field")
    write_table(ws3, ["field", *metric_headers], field_rows)
    ws4 = wb.create_sheet("By Document")
    write_table(ws4, ["doc_id", *metric_headers], doc_rows)

    detail_headers = [
        "doc_id", "field_group", "field", "field_kind",
        "gt_value", "pred_value", "status", "counts_as_tp",
        "matching_note", "char_similarity",
        "word_precision", "word_recall", "word_f1", "word_match_ratio", "word_jaccard",
        "matched_gt_row_id",
    ]
    ws5 = wb.create_sheet("Detail")
    write_table(ws5, detail_headers, detail_rows, status_col="status")
    ws6 = wb.create_sheet("Error Analysis")
    write_table(ws6, detail_headers, errors, status_col="status")
    ws7 = wb.create_sheet("Near Misses")
    write_table(ws7, detail_headers, near_misses, status_col="status")

    llm_rescued = [r for r in detail_rows if r.get("status") == "correct_llm"]
    if llm_rescued:
        ws_llm = wb.create_sheet("LLM Rescued")
        write_table(ws_llm, detail_headers, llm_rescued, status_col="status")

    if load_errors:
        ws8 = wb.create_sheet("Load Errors")
        write_table(ws8, ["doc_id", "error"], load_errors)
    if unmatched_predictions:
        ws9 = wb.create_sheet("Unmatched Predictions")
        write_table(ws9, ["prediction_file", "reason", "normalized_key", "candidate_count"], unmatched_predictions)

    wb.save(out_path)


# ── main evaluation loop ──────────────────────────────────────────────────────
def run_evaluation(
    gt_dir: Path,
    predictions: Path,
    out_path: Path,
    api_key: str | None,
    model: str,
) -> None:
    matches, unmatched_predictions = find_matches(predictions, gt_dir)
    records: list[dict[str, Any]] = []
    load_errors: list[dict[str, Any]] = []

    for i, match in enumerate(matches, 1):
        print(f"  [{i}/{len(matches)}] {match.doc_id}")
        try:
            gt_rows = rows_from_json(load_json(match.gt_path))
            pred_rows = rows_from_json(load_json(match.pred_path))
            records.extend(evaluate_document(match.doc_id, gt_rows, pred_rows, api_key, model))
        except Exception as exc:
            load_errors.append({"doc_id": match.doc_id, "error": str(exc)})

    overall = compute_prf1(records)
    summary: dict[str, Any] = {
        **overall,
        "n_documents": len(matches) - len(load_errors),
        "n_prediction_files": len(list(predictions.glob("*.json"))),
        "n_unmatched_predictions": len(unmatched_predictions),
        "n_load_errors": len(load_errors),
        "llm_model": model if api_key else "disabled",
        "llm_calls": _STATS.calls,
        "llm_cache_hits": _STATS.cache_hits,
        "llm_errors": _STATS.errors,
        "llm_rescues": _STATS.hits,
        "llm_tokens_in": _STATS.total_tokens_in,
        "llm_tokens_out": _STATS.total_tokens_out,
        "text_char_threshold": TEXT_CHAR_THRESHOLD,
        "text_jaccard_threshold": TEXT_JACCARD_THRESHOLD,
        "strict_field_kinds": "numeric, date, identifier",
        "fuzzy_field_kinds": "text",
        "llm_candidate_fields": ", ".join(sorted(LLM_CANDIDATE_FIELDS)),
    }

    errors, near_misses = error_analysis(records)
    write_excel_report(
        out_path,
        summary,
        breakdown(records, "field_group"),
        breakdown(records, "field"),
        breakdown(records, "doc_id"),
        records,
        errors,
        near_misses,
        load_errors,
        unmatched_predictions,
    )

    print(f"\nReport saved:        {out_path}")
    print(f"Documents evaluated: {summary['n_documents']}")
    print(f"Field accuracy:      {summary['field_accuracy']}")
    print(f"Exact accuracy:      {summary['exact_accuracy']}")
    print(f"Correct (exact):     {summary['correct_exact']}")
    print(f"Correct (fuzzy):     {summary['correct_fuzzy']}")
    print(f"Correct (LLM):       {summary['correct_llm']}")
    print(f"LLM calls:           {_STATS.calls}  (cache hits: {_STATS.cache_hits}, errors: {_STATS.errors})")
    if _STATS.total_tokens_in:
        print(f"LLM tokens:          {_STATS.total_tokens_in} in / {_STATS.total_tokens_out} out")


def main() -> None:
    parser = argparse.ArgumentParser(description="LLM-supervised invoice evaluation via OpenRouter")
    parser.add_argument(
        "--predictions",
        type=Path,
        default=Path("data/company_dataset_copy/present_runs/predictions_base_updated_test"),
        help="Directory of prediction JSON files",
    )
    parser.add_argument(
        "--gt",
        type=Path,
        default=Path("data/company_dataset_copy/test/GT"),
        help="Directory of ground-truth JSON files",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("data/eval/evaluation_report_llm.xlsx"),
        help="Output XLSX report path",
    )
    parser.add_argument(
        "--model",
        default=DEFAULT_MODEL,
        help=f"OpenRouter model string (default: {DEFAULT_MODEL})",
    )
    parser.add_argument(
        "--no-llm",
        action="store_true",
        help="Disable LLM calls; run heuristics only",
    )
    parser.add_argument(
        "--api-key",
        default=None,
        help="OpenRouter API key (overrides OPENROUTER_API_KEY env var)",
    )
    args = parser.parse_args()

    api_key: str | None = None
    if not args.no_llm:
        api_key = args.api_key or os.environ.get("OPENROUTER_API_KEY")
        if not api_key:
            print(
                "WARNING: OPENROUTER_API_KEY not set and --api-key not provided. "
                "Running in heuristic-only mode. "
                "Set the env var or pass --api-key to enable LLM supervision."
            )
        else:
            print(f"LLM supervision enabled: {args.model}")
            _STATS._cache = load_llm_cache()
    else:
        print("LLM supervision disabled (--no-llm).")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    run_evaluation(args.gt, args.predictions, args.out, api_key, args.model)

    if api_key:
        save_llm_cache(_STATS._cache)


if __name__ == "__main__":
    main()
